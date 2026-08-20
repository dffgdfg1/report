# -*- coding: utf-8 -*-
"""试验计划生成引擎：从「试验计划模板.xlsx」按测试项填出一张试验计划表。

模板结构（第一个 sheet「试验计划」）：
- 行1  标题「试验计划」（合并 C1:N1）
- 行2  表头信息：产品名称/型号、试验阶段、试验日期、计划编制人、计划核准、客户批准
- 行3~4 列标题（No./试验项目/标准号/试验标准/判定依据/样机数量/试验周期/
        计划时间(开始,结束)/实际时间(开始,结束)/试验结果/试验机构/备注）
- 行5  表格编号行（整行合并）
- 行6  分组行「第N组（样机号）」（整行合并）
- 行7  数据行样例

数据来自报告项目 project = {info:{...}, tests:[{...}]}。
"""
import os
import copy
import math
import openpyxl
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE = os.path.join(BASE, "模板库", "试验计划模板.xlsx")

# 数据行列 -> (取值函数)；1-based 列号
NCOLS = 14
DATA_ROW_TEMPLATE = 7   # 模板里数据行样例的行号
GROUP_ROW_TEMPLATE = 6  # 模板里分组行样例的行号
FORMNO_ROW = 5          # 表格编号行

# 试验标准/方法列(D=4)：嵌图相关尺寸估算
COND_COL0 = 3            # D 列 0-based
COND_IMG_MAX_W = 330     # 单张配图最大宽(px)，略小于 D 列宽
COND_IMG_MAX_H = 420     # 单张配图最大高(px)，避免竖图把行撑得过高
# D 列宽约 352px，14pt 宋体每个中文字≈18.7px、ASCII≈9.3px；
# 一行约放 18 个中文字 ≈ 37 个「字宽」。取 36 略偏保守(宁可行高一点，不让文字被图压住)。
COND_CHARS_PER_LINE = 36 # D 列一行约放的「字宽」(中文=2, ASCII=1)
COND_LINE_PX = 24        # 14pt 文本单行高(px，含行距)
COND_IMG_GAP_PX = 8      # 图与图/文字之间的间隙(px)
COND_PAD_TOP_PX = 4      # 单元格顶部留白
COND_PAD_BOTTOM_PX = 8   # 末图与单元格底边留白


def _sample_key(t):
    """取一个测试项的样机号作为分组键：优先「样机编号」字段，
    为空时回退到样品行(samples)首个编号，仍空则用空串。"""
    key = (t.get("sample_no", "") or "").strip()
    if key:
        return key
    samples = t.get("samples", []) or []
    for s in samples:
        no = (s.get("no", "") or "").strip()
        if no:
            # 多个样品行时用范围/顿号拼接，作为该测试的样机标识
            nos = [x.get("no", "").strip() for x in samples if x.get("no", "").strip()]
            return "、".join(nos) if len(nos) > 1 else no
    return ""


def _group_tests(tests):
    """按样机号分组，保持首次出现顺序。返回 [(sample_no, [test,...]), ...]。"""
    groups = []
    index = {}
    for t in tests:
        key = _sample_key(t)
        if key not in index:
            index[key] = len(groups)
            groups.append((key, []))
        groups[index[key]][1].append(t)
    return groups


def _text_px_height(text):
    """估算一段文字在 D 列渲染后的像素高度（按可视行数）。"""
    if not text:
        return 0
    lines = str(text).replace("\r\n", "\n").replace("\r", "\n").split("\n")
    total = 0
    for line in lines:
        width = sum(2 if ord(ch) >= 0x2E80 else 1 for ch in line)
        total += max(1, math.ceil(width / COND_CHARS_PER_LINE))
    return total * COND_LINE_PX


def _fmt_date(s):
    """日期原样输出（报告里已是 2026.08.14 点分格式）。"""
    return str(s or "").strip()


def _days_between(a, b):
    """由起止日期估算试验周期天数（含首尾）。解析失败返回空串。"""
    import re
    def parse(s):
        m = re.findall(r"\d+", str(s or ""))
        if len(m) >= 3:
            return tuple(int(x) for x in m[:3])
        return None
    pa, pb = parse(a), parse(b)
    if not pa or not pb:
        return ""
    import datetime
    try:
        da = datetime.date(*pa); db = datetime.date(*pb)
        n = (db - da).days + 1
        return f"{n}天" if n > 0 else ""
    except Exception:
        return ""


def _copy_style(src_cell, dst_cell):
    """复制单元格样式（字体/边框/对齐/填充/数字格式）。"""
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)


def _set_header(ws, info, applicant):
    """填第2行表头信息，保留原有「标签：」前缀，只把值接到冒号后。"""
    name = (info.get("sample_name", "") or "").strip()
    model = (info.get("sample_model", "") or "").strip()
    no = (info.get("sample_no", "") or "").strip()
    phase = (info.get("verify_phase", "") or "").strip()
    model_no = model + ("/" + no if no else "")

    def put(cell_ref, label, value):
        cell = ws[cell_ref]
        raw = str(cell.value or "")
        if not value:
            return  # 无值：保留模板原样（标签+空白）
        # 保留原首尾换行样式（模板里表头带前后换行做垂直留白）
        lead = "\n" if raw.startswith("\n") else ""
        tail = "\n" if raw.endswith("\n") else ""
        # 取「标签：」前缀（去掉原有换行后再统一补回，避免换行叠加）
        core = raw.strip()
        prefix = core
        for sep in ("：", ":"):
            if sep in core:
                prefix = core.split(sep)[0] + sep
                break
        cell.value = f"{lead}{prefix} {value}{tail}"

    put("A2", "产品名称/零件名称：", name)
    put("D2", "产品型号/零件号：", model_no)
    put("E2", "试验阶段：", phase)
    # F2 试验日期、I2 计划编制人、K2 计划核准：按需求留空（保留标签）
    put("M2", "客户批准：", applicant)


def _embed_cond_images(ws, row, text, imgs):
    """把试验条件配图竖排嵌进 D 列（row，1-based），文字在上、图在下、互不遮挡。
    返回该行需要的最小行高(pt)；无图返回 0。"""
    if not imgs:
        return 0
    try:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        from openpyxl.styles import Alignment
        from PIL import Image as PImage
    except Exception:
        return 0
    # 关键：让 D 列文字顶端对齐（模板默认垂直居中，行一高文字会跑到中间与图重叠）
    dcell = ws.cell(row, COND_COL0 + 1)
    al = dcell.alignment
    dcell.alignment = Alignment(horizontal=al.horizontal or "left", vertical="top",
                                wrap_text=True)
    # 图从文字下方开始竖排：顶部留白 + 文字高 + 间隙
    y = COND_PAD_TOP_PX + _text_px_height(text) + COND_IMG_GAP_PX
    placed = 0
    for im in imgs:
        fp = im.get("path", "")
        if not fp or not os.path.exists(fp):
            continue
        try:
            with PImage.open(fp) as pim:
                w0, h0 = pim.size
            if w0 <= 0 or h0 <= 0:
                continue
            # 先按宽度缩放，再按高度上限二次缩放，避免竖图过高
            scale = min(1.0, COND_IMG_MAX_W / float(w0))
            w = w0 * scale; h = h0 * scale
            if h > COND_IMG_MAX_H:
                scale2 = COND_IMG_MAX_H / h
                w *= scale2; h *= scale2
            w = int(w); h = int(h)
            xi = XLImage(fp)
            xi.width = w; xi.height = h
            frm = AnchorMarker(col=COND_COL0, colOff=pixels_to_EMU(4),
                               row=row - 1, rowOff=pixels_to_EMU(y))
            size = XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h))
            xi.anchor = OneCellAnchor(_from=frm, ext=size)
            ws.add_image(xi)
            y += h + COND_IMG_GAP_PX
            placed += 1
        except Exception:
            continue
    if placed == 0:
        return 0
    y += COND_PAD_BOTTOM_PX
    # px -> pt (行高单位)：1px ≈ 0.75pt
    return y * 0.75


def _write_data_rows(ws, groups):
    """把分组后的 tests 写进数据区。groups = [(sample_no, [test,...]), ...]。
    分组行样例(行6)+数据行样例(行7)作为样式来源。"""
    grp_styles = [ws.cell(GROUP_ROW_TEMPLATE, c) for c in range(1, NCOLS + 1)]
    dat_styles = [ws.cell(DATA_ROW_TEMPLATE, c) for c in range(1, NCOLS + 1)]
    grp_h = ws.row_dimensions[GROUP_ROW_TEMPLATE].height

    # 清掉模板样例行里自带的配图（否则会与新图叠加）
    try:
        ws._images = []
    except Exception:
        pass
    # 清掉模板里第6行起的合并
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= GROUP_ROW_TEMPLATE:
            ws.unmerge_cells(str(mc))
    ws.delete_rows(GROUP_ROW_TEMPLATE, 2)  # 删样例行(6,7)，从第6行重建

    # 清掉数据样例以下遗留的空行内容
    for rr in range(GROUP_ROW_TEMPLATE, ws.max_row + 1):
        for cc in range(1, NCOLS + 1):
            ws.cell(rr, cc).value = None

    r = GROUP_ROW_TEMPLATE
    seq = 1
    for gi, (sample_no, gtests) in enumerate(groups, 1):
        for c in range(1, NCOLS + 1):
            _copy_style(grp_styles[c - 1], ws.cell(r, c))
        ws.cell(r, 1).value = f"第{gi}组（{sample_no}）"
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        if grp_h:
            ws.row_dimensions[r].height = grp_h
        r += 1
        for t in gtests:
            vals = _row_values(t, seq)
            for c in range(1, NCOLS + 1):
                cell = ws.cell(r, c)
                _copy_style(dat_styles[c - 1], cell)
                cell.value = vals[c - 1]
            # 试验条件配图嵌入 D 列，并据此撑高行
            cond_text = t.get("condition", "")
            cimgs = t.get("condition_images", []) or []
            need_h = _embed_cond_images(ws, r, cond_text, cimgs)
            if need_h > 0:
                ws.row_dimensions[r].height = max(need_h, 30)
            r += 1
            seq += 1


def _row_values(t, seq):
    """返回一行 14 列的值列表。"""
    samples = t.get("samples", []) or []
    qty = len(samples) if samples else ""
    start = _fmt_date(t.get("start_date", ""))
    end = _fmt_date(t.get("end_date", ""))
    cycle = _days_between(start, end)
    return [
        seq,                                  # 1 No.
        t.get("title", ""),                   # 2 试验项目
        t.get("standard", ""),                # 3 标准号/条款号
        t.get("condition", ""),               # 4 试验标准/方法
        t.get("requirement", ""),             # 5 判定依据
        qty,                                  # 6 样机数量
        cycle,                                # 7 试验周期
        "",                                   # 8 计划时间-开始（留空）
        "",                                   # 9 计划时间-结束（留空）
        start,                                # 10 实际时间-开始
        end,                                  # 11 实际时间-结束
        t.get("overall_result", ""),          # 12 试验结果
        "minieye",                            # 13 试验机构
        "",                                   # 14 备注
    ]


# 分组表布局：组横排在 A/C/E/G/I/K 列（0 间隔 B/D/F/H/J），最多 6 组
GRP_SHEET_COLS = [1, 3, 5, 7, 9, 11]   # 1-based：A C E G I K
GRP_SHEET_HDR_ROW = 5                   # 组标题行
GRP_SHEET_ITEM_ROW = 6                  # 第一条测试项行


def _fill_group_sheet(ws, groups):
    """填「分组表」sheet：每组一列，标题「第N组（样机号）」下竖排该组各试验项目名。"""
    # 样式来源：标题格(A5)、项目格(A6)
    hdr_style = ws.cell(GRP_SHEET_HDR_ROW, 1)
    item_style = ws.cell(GRP_SHEET_ITEM_ROW, 1)
    hdr_h = ws.row_dimensions[GRP_SHEET_HDR_ROW].height or 21.95
    item_h = ws.row_dimensions[GRP_SHEET_ITEM_ROW].height or 30.0

    # 先清空模板里 6 组样例的标题与项目格内容（保留样式来源前先取好）
    from copy import copy as _c
    hstyle = {"font": _c(hdr_style.font), "border": _c(hdr_style.border),
              "fill": _c(hdr_style.fill), "alignment": _c(hdr_style.alignment)}
    istyle = {"font": _c(item_style.font), "border": _c(item_style.border),
              "fill": _c(item_style.fill), "alignment": _c(item_style.alignment)}

    def apply(cell, st):
        cell.font = _c(st["font"]); cell.border = _c(st["border"])
        cell.fill = _c(st["fill"]); cell.alignment = _c(st["alignment"])

    # 只填有数据的组；模板里其余空组（第N组（）+空框）原样保留，不删除。
    for gi, (sample_no, gtests) in enumerate(groups):
        if gi >= len(GRP_SHEET_COLS):
            break  # 模板仅 6 组位
        col = GRP_SHEET_COLS[gi]
        # 覆盖该组标题（沿用绿底红字样式）
        hc = ws.cell(GRP_SHEET_HDR_ROW, col)
        apply(hc, hstyle)
        hc.value = f"第{gi+1}组（{sample_no}）"
        # 竖排项目：每项之间空一行（占偶数偏移行）。
        # 超出模板自带空框的行(第10行起)才补边框；不动其它组的列。
        for k, t in enumerate(gtests):
            row = GRP_SHEET_ITEM_ROW + k * 2
            ic = ws.cell(row, col)
            apply(ic, istyle)
            ic.value = t.get("title", "")
            if ws.row_dimensions[row].height is None:
                ws.row_dimensions[row].height = item_h


def generate_plan(project, out_path):
    """project: {info:{...}, tests:[...]} -> 生成试验计划 .xlsx，返回文件路径。"""
    info = project.get("info", {}) or {}
    tests = project.get("tests", []) or []
    applicant = (info.get("applicant", "") or "").strip()

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["试验计划"] if "试验计划" in wb.sheetnames else wb.active

    groups = _group_tests(tests)
    _set_header(ws, info, applicant)
    if tests:
        _write_data_rows(ws, groups)

    # 第二个 sheet「分组表」
    if "分组表" in wb.sheetnames and groups:
        _fill_group_sheet(wb["分组表"], groups)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path
