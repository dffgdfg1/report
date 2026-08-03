# -*- coding: utf-8 -*-
"""本地报告生成器 Web 后端。仅监听 127.0.0.1，不联网。"""
import os, re, json, uuid, glob, webbrowser, threading, sys, subprocess
from flask import Flask, request, jsonify, send_file, send_from_directory, abort

def _is_local_request():
    """请求是否来自运行服务器的本机（localhost）。"""
    return (request.remote_addr or "") in ("127.0.0.1", "::1", "localhost")

def _native_open(path):
    """在服务器本机用系统默认程序打开文件/文件夹（跨平台）。"""
    if sys.platform.startswith("win"):
        os.startfile(path)  # noqa: Windows 专用，仅在 Windows 上调用
    elif sys.platform == "darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])

import presets as P
import report_engine as E

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROJ_DIR = os.path.join(BASE, "项目")
OUT_DIR = os.path.join(BASE, "输出")
os.makedirs(PROJ_DIR, exist_ok=True)
os.makedirs(OUT_DIR, exist_ok=True)

app = Flask(__name__, static_folder="static", template_folder="templates")
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 单次上传上限 200MB

SAFE = re.compile(r"[^0-9A-Za-z_\-一-鿿]")
def safe_name(s):
    s = SAFE.sub("_", (s or "").strip())
    return s or "未命名"

def pdir(name):
    d = os.path.join(PROJ_DIR, safe_name(name))
    os.makedirs(os.path.join(d, "images"), exist_ok=True)
    return d

@app.route("/")
def index():
    return send_from_directory("templates", "index.html")

@app.after_request
def _cache_policy(resp):
    """代码文件(html/js/css)禁缓存，避免改动不生效；
    图片(uuid命名、内容不变)长期缓存，避免每次刷新缩略图都重新拉取、压垮服务器。"""
    p = request.path or ""
    if p.startswith("/api/image/"):
        resp.headers["Cache-Control"] = "public, max-age=604800"  # 图片缓存一周
    elif p.endswith((".js", ".css")) or p == "/" or p.endswith(".html"):
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
    return resp

@app.route("/api/health")
def api_health():
    return jsonify({"ok": True})

@app.route("/api/meta")
def api_meta():
    return jsonify({"types": P.list_types(),
                    "presets": {k: P.get_preset(k) for k in P.list_types()},
                    "img_w": E.IMG_W_CM, "img_h": E.IMG_H_CM, "fit": E.FIT_MODE})

@app.route("/api/projects")
def api_projects():
    names = []
    for d in sorted(glob.glob(os.path.join(PROJ_DIR, "*"))):
        if os.path.isdir(d) and os.path.exists(os.path.join(d, "project.json")):
            names.append(os.path.basename(d))
    return jsonify(names)

@app.route("/api/load")
def api_load():
    name = request.args.get("name", "")
    fp = os.path.join(pdir(name), "project.json")
    if not os.path.exists(fp):
        return jsonify({"info": {}, "tests": [], "name": name})
    with open(fp, "r", encoding="utf-8") as f:
        return jsonify(json.load(f))

def _backup_project(fp):
    """保存前把旧 project.json 复制成带时间戳的备份，保留最近 10 份。"""
    if not os.path.exists(fp):
        return
    import shutil, datetime
    d = os.path.dirname(fp)
    bdir = os.path.join(d, "备份")
    os.makedirs(bdir, exist_ok=True)
    # 毫秒级时间戳，避免同一秒内多次保存互相覆盖
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
    try:
        shutil.copy2(fp, os.path.join(bdir, "project.%s.json" % ts))
    except Exception:
        return
    # 只保留最近 10 份
    baks = sorted(glob.glob(os.path.join(bdir, "project.*.json")))
    for old in baks[:-10]:
        try:
            os.remove(old)
        except Exception:
            pass

@app.route("/api/save", methods=["POST"])
def api_save():
    data = request.get_json(force=True)
    name = data.get("name") or data.get("info", {}).get("report_no") or "未命名"
    data["name"] = name
    fp = os.path.join(pdir(name), "project.json")
    _backup_project(fp)  # 覆盖前先备份
    with open(fp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return jsonify({"ok": True, "name": name})

# ---------- 测试方案库（用户自存的测试项目模板） ----------
SCHEME_FILE = os.path.join(BASE, "模板库", "测试方案.json")

def _load_schemes():
    if os.path.exists(SCHEME_FILE):
        try:
            with open(SCHEME_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def _save_schemes(d):
    os.makedirs(os.path.dirname(SCHEME_FILE), exist_ok=True)
    with open(SCHEME_FILE, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)

# 方案里只保留"可复用"的字段，不含图片/日期等一次性内容
SCHEME_KEYS = ["title", "standard", "env", "condition", "requirement",
               "overall_result", "equipment", "image_group_titles"]

@app.route("/api/schemes")
def api_schemes():
    return jsonify(sorted(_load_schemes().keys()))

@app.route("/api/scheme/save", methods=["POST"])
def api_scheme_save():
    data = request.get_json(force=True)
    sname = (data.get("scheme_name") or "").strip()
    test = data.get("test", {}) or {}
    if not sname:
        return jsonify({"ok": False, "error": "缺少方案名称"}), 400
    scheme = {k: test.get(k) for k in SCHEME_KEYS if test.get(k) is not None}
    # 图组标题从 image_groups 提取（保存分组名，不存图片）
    if test.get("image_groups"):
        scheme["image_group_titles"] = [g.get("title", "") for g in test["image_groups"]]
    schemes = _load_schemes()
    schemes[sname] = scheme
    _save_schemes(schemes)
    return jsonify({"ok": True, "name": sname, "schemes": sorted(schemes.keys())})

@app.route("/api/scheme/get")
def api_scheme_get():
    sname = request.args.get("name", "")
    schemes = _load_schemes()
    if sname not in schemes:
        return jsonify({"ok": False, "error": "方案不存在"}), 404
    return jsonify({"ok": True, "scheme": schemes[sname]})

@app.route("/api/scheme/delete", methods=["POST"])
def api_scheme_delete():
    data = request.get_json(force=True)
    sname = (data.get("name") or "").strip()
    schemes = _load_schemes()
    if sname in schemes:
        del schemes[sname]
        _save_schemes(schemes)
    return jsonify({"ok": True, "schemes": sorted(schemes.keys())})

# ---------- 标准条件库（测试项目 × 车厂 → 条件） ----------
STD_FILE = os.path.join(BASE, "模板库", "标准库.json")
STD_IMG_DIR = os.path.join(BASE, "模板库", "std_images")  # 标准附图（PSD谱/曲线/表格截图等）
os.makedirs(STD_IMG_DIR, exist_ok=True)

def _load_std():
    if os.path.exists(STD_FILE):
        try:
            with open(STD_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def _save_std(d):
    os.makedirs(os.path.dirname(STD_FILE), exist_ok=True)
    with open(STD_FILE, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)

@app.route("/api/standards")
def api_standards():
    """返回整个标准库：{测试项目: {车厂: {standard, condition, requirement}}}"""
    return jsonify(_load_std())

# Excel 表头名 -> 内部字段名
_STD_COLMAP = {
    "测试项目": "item", "车厂": "oem",
    "标准号/条款号": "standard", "标准号": "standard",
    "试验条件": "condition", "试验要求": "requirement",
    "备注": "remark",
}

@app.route("/api/standards/import", methods=["POST"])
def api_standards_import():
    """上传填好的 Excel 模板，解析并合并进标准库。"""
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "未收到文件"}), 400
    try:
        from openpyxl import load_workbook
    except Exception:
        return jsonify({"ok": False, "error": "缺少 openpyxl 组件，无法读取 Excel。请联系维护者。"}), 200
    import io
    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as ex:
        return jsonify({"ok": False, "error": "打不开这个 Excel：%s" % ex}), 200
    # 优先「标准条件」页，否则用第一个页
    ws = wb["标准条件"] if "标准条件" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"ok": False, "error": "表格是空的"}), 200
    # 表头 -> 列索引
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {}
    for i, h in enumerate(header):
        if h in _STD_COLMAP:
            idx[_STD_COLMAP[h]] = i
    if "item" not in idx or "oem" not in idx:
        return jsonify({"ok": False, "error": "表头不对，缺「测试项目」或「车厂」列。请用最新模板。"}), 200

    def cell(row, key):
        i = idx.get(key)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        return "" if v is None else str(v).strip()

    std = _load_std()
    added, updated, skipped = 0, 0, 0
    for row in rows[1:]:
        if row is None:
            continue
        item = cell(row, "item")
        oem = cell(row, "oem")
        # 跳过空行 / 示例行
        if not item or not oem:
            skipped += 1
            continue
        if "示例" in cell(row, "remark"):
            skipped += 1
            continue
        entry = {
            "standard": cell(row, "standard"),
            "condition": cell(row, "condition"),
            "requirement": cell(row, "requirement"),
        }
        if item not in std:
            std[item] = {}
        if oem in std[item]:
            # 覆盖文本字段，但保留已有附图
            prev = std[item][oem]
            if isinstance(prev, dict) and prev.get("images"):
                entry["images"] = prev["images"]
            updated += 1
        else:
            added += 1
        std[item][oem] = entry

    # 提取 Excel 里的浮动图片，按锚点行归属到对应条目
    imgs_imported, imgs_orphan = 0, 0
    _ext_map = {"jpeg": "jpg", "jpg": "jpg", "png": "png", "gif": "gif", "bmp": "bmp"}
    try:
        sheet_imgs = list(getattr(ws, "_images", []) or [])
    except Exception:
        sheet_imgs = []

    def _anchor_rc(im):
        a = getattr(getattr(im, "anchor", None), "_from", None)
        return (getattr(a, "row", 10 ** 9), getattr(a, "col", 10 ** 9))
    # 同一行的多张图按列从左到右排序，保证多图顺序稳定
    sheet_imgs.sort(key=_anchor_rc)

    for im in sheet_imgs:
        try:
            r0, _c0 = _anchor_rc(im)  # 0-based；只按行归属，列只用来排序
            if r0 is None or r0 <= 0 or r0 >= len(rows):
                imgs_orphan += 1
                continue
            row = rows[r0]
            item = cell(row, "item"); oem = cell(row, "oem")
            if not item or not oem or item not in std or oem not in std[item]:
                imgs_orphan += 1
                continue
            data = None
            try:
                data = im._data()
            except Exception:
                data = None
            if not data:
                imgs_orphan += 1
                continue
            fmt = str(getattr(im, "format", "") or "png").lower()
            ext = _ext_map.get(fmt, "png")
            fn = "%s.%s" % (uuid.uuid4().hex, ext)
            with open(os.path.join(STD_IMG_DIR, fn), "wb") as wf:
                wf.write(data)
            tgt = std[item][oem]
            tgt.setdefault("images", []).append(
                {"file": fn, "orig": "Excel嵌入图.%s" % ext, "caption": ""})
            imgs_imported += 1
        except Exception:
            imgs_orphan += 1

    _save_std(std)
    items = len(std)
    pairs = sum(len(v) for v in std.values())
    return jsonify({"ok": True, "added": added, "updated": updated,
                    "skipped": skipped, "items": items, "pairs": pairs,
                    "images": imgs_imported, "images_orphan": imgs_orphan})

@app.route("/api/standards/clear", methods=["POST"])
def api_standards_clear():
    """清空整个标准库（谨慎）。"""
    _save_std({})
    return jsonify({"ok": True})

@app.route("/api/standards/upsert", methods=["POST"])
def api_standards_upsert():
    """新增或修改一条标准。body:
    {old_item, old_oem, item, oem, standard, condition, requirement}
    old_* 为编辑前的键；若与新键不同则先删旧键（支持改名/改车厂）。"""
    d = request.get_json(force=True) or {}
    item = (d.get("item") or "").strip()
    oem = (d.get("oem") or "").strip()
    if not item or not oem:
        return jsonify({"ok": False, "error": "「测试项目」和「车厂」都不能为空"}), 200
    old_item = (d.get("old_item") or "").strip()
    old_oem = (d.get("old_oem") or "").strip()
    std = _load_std()
    # 编辑时若键变了，先删旧的
    if old_item and old_oem and (old_item != item or old_oem != oem):
        if old_item in std and old_oem in std[old_item]:
            del std[old_item][old_oem]
            if not std[old_item]:
                del std[old_item]
    is_new = not (item in std and oem in std.get(item, {}))
    # 保留原有附图（编辑时可能改了键，附图从旧键/新键任一处继承）
    prev = {}
    if old_item and old_oem:
        prev = ((std.get(old_item) or {}).get(old_oem)) or {}
    if not prev:
        prev = ((std.get(item) or {}).get(oem)) or {}
    std.setdefault(item, {})[oem] = {
        "standard": (d.get("standard") or "").strip(),
        "condition": (d.get("condition") or "").strip(),
        "requirement": (d.get("requirement") or "").strip(),
        "images": prev.get("images", []),
    }
    _save_std(std)
    return jsonify({"ok": True, "created": is_new,
                    "items": len(std), "pairs": sum(len(v) for v in std.values())})

@app.route("/api/standards/delete", methods=["POST"])
def api_standards_delete():
    """删除一条(给 item+oem)或整个测试项目(只给 item)。"""
    d = request.get_json(force=True) or {}
    item = (d.get("item") or "").strip()
    oem = (d.get("oem") or "").strip()
    if not item:
        return jsonify({"ok": False, "error": "缺少测试项目"}), 200
    std = _load_std()
    if item not in std:
        return jsonify({"ok": True, "items": len(std), "pairs": sum(len(v) for v in std.values())})
    if oem:
        std[item].pop(oem, None)
        if not std[item]:
            del std[item]
    else:
        del std[item]
    _save_std(std)
    return jsonify({"ok": True, "items": len(std), "pairs": sum(len(v) for v in std.values())})

@app.route("/api/standards/template")
def api_standards_template():
    """下载标准库 Excel 填写模板（含「附图」列与填写说明）。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception:
        # 兜底：回退到静态模板文件
        fp = os.path.join(BASE, "模板库", "标准库模板.xlsx")
        if os.path.exists(fp):
            return send_file(fp, as_attachment=True, download_name="标准库模板.xlsx")
        return jsonify({"ok": False, "error": "缺少 openpyxl 组件，无法生成模板。"}), 200
    import io
    from openpyxl.utils import get_column_letter
    IMG_COLS = 6   # 预留 6 个附图列，一格一张，够放多张图
    wb = Workbook()
    ws = wb.active
    ws.title = "标准条件"
    TEXT_COLS = ["测试项目", "车厂", "标准号/条款号", "试验条件", "试验要求", "备注"]
    img_headers = ["附图%d" % i for i in range(1, IMG_COLS + 1)]
    headers = TEXT_COLS + img_headers
    ncols = len(headers)
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="D9E4F5")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(vertical="center")
    widths = [22, 12, 22, 30, 26, 14] + [28] * IMG_COLS
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # 示例行 + 说明（备注含「示例」会在导入时自动跳过）
    ws.append(["振动测试", "示例车厂", "GB/T 2423.10", "10-500Hz 扫频 2h/轴",
               "无结构损坏、功能正常", "示例（导入时自动跳过）"] + [""] * IMG_COLS)
    note = ("填写说明：每行一条标准（测试项目 + 车厂 唯一）。"
            "如有附图（曲线/谱表/表格截图），把图片插入到本行后面的「附图1、附图2…」列里，"
            "一个格子放一张，从附图1往后依次放，图片够放多张——导入时会自动全部归到该行对应的标准。"
            "备注里含「示例」二字的行会被跳过。")
    ws.append([note])
    try:
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ncols)
        nc = ws.cell(row=ws.max_row, column=1)
        nc.font = Font(color="888888", italic=True)
        nc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ws.max_row].height = 56
    except Exception:
        pass
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="标准库模板.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _std_export_embed_images(ws, excel_row, first_img_col0, images, ok, fail):
    """把一条标准的附图逐张嵌进「附图1、附图2…」各列（一格一张），
    锚定到 excel_row（1-based）。行高取本行最高的图，保证再次导入时按行归位。"""
    if not images:
        return ok, fail
    try:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        from PIL import Image as PImage
    except Exception:
        return ok, fail
    MAXW = 200  # 单格缩略宽度上限(px)
    col = first_img_col0
    max_h = 0
    for rec in images:
        try:
            fp = os.path.join(STD_IMG_DIR, rec.get("file", ""))
            if not rec.get("file") or not os.path.exists(fp):
                fail += 1
                continue
            with PImage.open(fp) as pim:
                w0, h0 = pim.size
            if w0 <= 0 or h0 <= 0:
                fail += 1
                continue
            scale = min(1.0, MAXW / float(w0))
            w = int(w0 * scale); h = int(h0 * scale)
            xi = XLImage(fp)
            xi.width = w; xi.height = h
            frm = AnchorMarker(col=col, colOff=pixels_to_EMU(4),
                               row=excel_row - 1, rowOff=pixels_to_EMU(2))
            size = XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h))
            xi.anchor = OneCellAnchor(_from=frm, ext=size)
            ws.add_image(xi)
            col += 1              # 下一张放到右边一列
            max_h = max(max_h, h)
            ok += 1
        except Exception:
            fail += 1
    if max_h:
        # 行高单位是 pt，1px≈0.75pt；留点余量
        try:
            ws.row_dimensions[excel_row].height = max(18, max_h * 0.75 + 6)
        except Exception:
            pass
    return ok, fail

@app.route("/api/standards/export")
def api_standards_export():
    """把当前标准库导出为 Excel（表头与导入模板一致，可再次导入）。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception:
        return jsonify({"ok": False, "error": "缺少 openpyxl 组件，无法导出。"}), 200
    import io, datetime
    from openpyxl.utils import get_column_letter
    std = _load_std()
    # 每条标准最多几张图 -> 需要几列附图（至少留 1 列）
    max_imgs = 1
    for item in std.values():
        for e in item.values():
            n = len((e or {}).get("images", []) or [])
            if n > max_imgs:
                max_imgs = n
    TEXT_COLS = ["测试项目", "车厂", "标准号/条款号", "试验条件", "试验要求", "备注"]
    img_headers = ["附图%d" % i for i in range(1, max_imgs + 1)]
    headers = TEXT_COLS + img_headers
    wb = Workbook()
    ws = wb.active
    ws.title = "标准条件"
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="D9E4F5")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(vertical="center")
    widths = [22, 12, 22, 30, 26, 14] + [28] * len(img_headers)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    FIRST_IMG_COL0 = len(TEXT_COLS)   # 0-based：第一张附图所在列
    blank_imgs = [""] * len(img_headers)
    r = 1                             # 当前 excel 行（表头是第1行）
    embed_ok, embed_fail = 0, 0
    for item in sorted(std.keys()):
        for oem in sorted(std[item].keys()):
            e = std[item][oem] or {}
            ws.append([item, oem, e.get("standard", ""), e.get("condition", ""),
                       e.get("requirement", ""), ""] + blank_imgs)
            r += 1
            embed_ok, embed_fail = _std_export_embed_images(
                ws, r, FIRST_IMG_COL0, e.get("images", []), embed_ok, embed_fail)
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fn = "标准库备份_%s.xlsx" % datetime.date.today().strftime("%Y%m%d")
    return send_file(bio, as_attachment=True, download_name=fn,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---- 标准附图：给某条标准(item×oem)挂图片(PSD谱/曲线/表格截图等) ----
def _std_entry(std, item, oem):
    return (std.get(item) or {}).get(oem)

@app.route("/api/standards/image/upload", methods=["POST"])
def api_std_image_upload():
    """给一条标准上传附图。form: item, oem, files[]。"""
    item = (request.form.get("item") or "").strip()
    oem = (request.form.get("oem") or "").strip()
    if not item or not oem:
        return jsonify({"ok": False, "error": "缺少 item / oem"}), 400
    std = _load_std()
    e = _std_entry(std, item, oem)
    if e is None:
        return jsonify({"ok": False, "error": "这条标准不存在，请先保存基本信息再传图。"}), 200
    OK_EXT = (".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp", ".tif", ".tiff")
    e.setdefault("images", [])
    saved = []
    for f in request.files.getlist("files"):
        ext = os.path.splitext(f.filename)[1].lower() or ".jpg"
        if ext not in OK_EXT:
            continue
        fn = uuid.uuid4().hex + ext
        f.save(os.path.join(STD_IMG_DIR, fn))
        rec = {"file": fn, "orig": f.filename, "caption": ""}
        e["images"].append(rec)
        saved.append(rec)
    _save_std(std)
    return jsonify({"ok": True, "files": saved, "count": len(e["images"])})

@app.route("/api/standards/image/<fn>")
def api_std_image(fn):
    if not os.path.exists(os.path.join(STD_IMG_DIR, fn)):
        abort(404)
    return send_from_directory(STD_IMG_DIR, fn)

@app.route("/api/standards/image/delete", methods=["POST"])
def api_std_image_delete():
    """删除一条标准里的某张附图。body: item, oem, file。"""
    d = request.get_json(force=True) or {}
    item, oem, file = (d.get("item") or "").strip(), (d.get("oem") or "").strip(), (d.get("file") or "")
    std = _load_std()
    e = _std_entry(std, item, oem)
    if e is None:
        return jsonify({"ok": False, "error": "标准不存在"}), 200
    e["images"] = [im for im in e.get("images", []) if im.get("file") != file]
    _save_std(std)
    # 若没有其它标准引用这张图，删除物理文件
    used = any(im.get("file") == file
               for it in std.values() for en in it.values() for im in en.get("images", []))
    if not used:
        try:
            os.remove(os.path.join(STD_IMG_DIR, file))
        except Exception:
            pass
    return jsonify({"ok": True, "count": len(e["images"])})

@app.route("/api/standards/image/caption", methods=["POST"])
def api_std_image_caption():
    """改附图的图注。body: item, oem, file, caption。"""
    d = request.get_json(force=True) or {}
    item, oem, file = (d.get("item") or "").strip(), (d.get("oem") or "").strip(), (d.get("file") or "")
    std = _load_std()
    e = _std_entry(std, item, oem)
    if e is None:
        return jsonify({"ok": False, "error": "标准不存在"}), 200
    for im in e.get("images", []):
        if im.get("file") == file:
            im["caption"] = d.get("caption", "")
    _save_std(std)
    return jsonify({"ok": True})

@app.route("/api/standards/image/apply", methods=["POST"])
def api_std_image_apply():
    """套用标准时，把该标准的附图复制进当前项目的 images 目录，
    返回可直接写进 test.condition_images 的记录列表。body: project, item, oem。"""
    import shutil
    d = request.get_json(force=True) or {}
    project = (d.get("project") or "").strip()
    item, oem = (d.get("item") or "").strip(), (d.get("oem") or "").strip()
    if not project:
        return jsonify({"ok": False, "error": "缺少项目名"}), 400
    std = _load_std()
    e = _std_entry(std, item, oem)
    imgs = (e or {}).get("images", []) if e else []
    dst_dir = os.path.join(pdir(project), "images")
    out = []
    for im in imgs:
        src = os.path.join(STD_IMG_DIR, im.get("file", ""))
        if not os.path.exists(src):
            continue
        ext = os.path.splitext(im["file"])[1] or ".jpg"
        newfn = uuid.uuid4().hex + ext
        try:
            shutil.copyfile(src, os.path.join(dst_dir, newfn))
        except Exception:
            continue
        out.append({"file": newfn, "orig": im.get("orig", ""), "caption": im.get("caption", ""), "from_std": True})
    return jsonify({"ok": True, "images": out})

# 生成日志：记录每次实际生成的报告编号，用于报告编号自动递增
GENLOG_FILE = os.path.join(BASE, "输出", "生成记录.json")

def _load_genlog():
    if os.path.exists(GENLOG_FILE):
        try:
            with open(GENLOG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def _append_genlog(report_no):
    if not report_no:
        return
    import datetime
    log = _load_genlog()
    log.append({"report_no": report_no, "at": datetime.datetime.now().isoformat(timespec="seconds")})
    log = log[-2000:]  # 只留最近 2000 条，防止无限增长
    try:
        with open(GENLOG_FILE, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# ============ 设备库：从实验室设备清单 CSV 建库，报告里手动选设备 ============
DEV_FILE = os.path.join(BASE, "模板库", "设备库.json")

def _load_dev():
    if os.path.exists(DEV_FILE):
        try:
            with open(DEV_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def _save_dev(d):
    os.makedirs(os.path.dirname(DEV_FILE), exist_ok=True)
    with open(DEV_FILE, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)

def _dev_cal_end(cal_date):
    """校准有效期止 = 校准日期 + 1年 - 1天。入参/出参都用 YYYY.MM.DD。"""
    import datetime, re
    m = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", cal_date or "")
    if not m:
        return ""
    y, mo, da = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        start = datetime.date(y, mo, da)
    except ValueError:
        return ""
    try:
        nxt = start.replace(year=y + 1)
    except ValueError:            # 2/29 -> 次年用 2/28
        nxt = start.replace(year=y + 1, day=28)
    end = nxt - datetime.timedelta(days=1)
    return "%04d.%02d.%02d" % (end.year, end.month, end.day)

def _dot_date(s):
    import re
    m = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", s or "")
    return "%04d.%02d.%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else ""

def _dev_clean_model(s):
    """型号里去掉中文：型号是纯 ASCII 编码，中文都在单独行（是描述），整行丢弃。
    例：'三综合试验箱\\nXB-OTS-408F-B' -> 'XB-OTS-408F-B'；'振动试验台\\n（1吨）\\nDC-1000-15' -> 'DC-1000-15'。"""
    s = (s or "").strip()
    if not s or s == "/":
        return ""
    kept = []
    for ln in re.split(r"[\r\n]+", s):
        ln = ln.strip()
        if not ln:
            continue
        if re.search(r"[㐀-鿿]", ln):   # 含中文汉字 -> 描述行，丢弃
            continue
        kept.append(ln)
    out = " ".join(kept).strip()
    return "" if out == "/" else out

def _dev_key(d):
    """设备唯一键：管理编号优先，否则名称|型号|出厂编号。"""
    return (d.get("mgmt_no") or "").strip() or \
        "%s|%s|%s" % (d.get("name", ""), d.get("model", ""), d.get("factory_no", ""))

@app.route("/api/devices")
def api_devices():
    """返回整个设备库：[{name, model, mgmt_no, cal_date, cal_end, factory_no, maker, remark}]"""
    return jsonify(_load_dev())

# CSV 表头名 -> 内部字段（兼容全/半角空格）
_DEV_COLMAP = {
    "名称": "name", "名  称": "name", "名 称": "name",
    "型号": "model",
    "设备编号": "mgmt_no",
    "校准日期": "cal_date",
    "出厂编号": "factory_no",
    "设备厂家": "maker",
    "测试类型": "test_type",
    "备注": "remark",
}

def _dev_norm_header(h):
    return re.sub(r"\s+", "", str(h or "")).strip()

@app.route("/api/devices/import", methods=["POST"])
def api_devices_import():
    """上传实验室设备清单 CSV，解析建库（覆盖式重建）。"""
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "未收到文件"}), 400
    import csv, io as _io
    raw = f.read()
    txt = None
    for enc in ("utf-8-sig", "gbk", "utf-8", "gb18030"):
        try:
            txt = raw.decode(enc)
            break
        except Exception:
            continue
    if txt is None:
        return jsonify({"ok": False, "error": "无法识别文件编码，请用 UTF-8 或 GBK 的 CSV。"}), 200
    rows = list(csv.reader(_io.StringIO(txt)))
    if not rows:
        return jsonify({"ok": False, "error": "表格是空的"}), 200
    # 找表头行：含「名称」列且含「校准日期」列
    hidx = -1
    for i, r in enumerate(rows[:5]):
        norm = [_dev_norm_header(c) for c in r]
        if any(n in ("名称",) for n in norm) and "校准日期" in norm:
            hidx = i
            break
    if hidx < 0:
        hidx = 0
    header = [_dev_norm_header(c) for c in rows[hidx]]
    idx = {}
    for i, h in enumerate(header):
        key = _DEV_COLMAP.get(h)
        if key and key not in idx:
            idx[key] = i
    if "name" not in idx:
        return jsonify({"ok": False, "error": "表头里找不到「名称」列，请确认是设备清单 CSV。"}), 200

    def cell(row, key):
        i = idx.get(key)
        if i is None or i >= len(row):
            return ""
        return (row[i] or "").strip()

    devices = []
    seen = set()
    for row in rows[hidx + 1:]:
        if not row or not any((c or "").strip() for c in row):
            continue
        name = cell(row, "name")
        if not name or name == "/":
            continue
        cal_date = _dot_date(cell(row, "cal_date"))
        rec = {
            "name": name,
            "model": _dev_clean_model(cell(row, "model")),
            "mgmt_no": cell(row, "mgmt_no"),
            "cal_date": cal_date,
            "cal_end": _dev_cal_end(cal_date),
            "factory_no": cell(row, "factory_no"),
            "maker": cell(row, "maker"),
            "test_type": cell(row, "test_type").replace("\n", " ").strip(),
            "remark": cell(row, "remark"),
        }
        # 去重键：设备编号优先，否则名称+型号+出厂编号
        k = rec["mgmt_no"] or "%s|%s|%s" % (rec["name"], rec["model"], rec["factory_no"])
        if k in seen:
            continue
        seen.add(k)
        devices.append(rec)
    devices.sort(key=lambda d: (d.get("name", ""), d.get("mgmt_no", "")))
    _save_dev(devices)
    return jsonify({"ok": True, "count": len(devices)})

@app.route("/api/devices/clear", methods=["POST"])
def api_devices_clear():
    _save_dev([])
    return jsonify({"ok": True})

@app.route("/api/devices/export")
def api_devices_export():
    """导出设备库为 CSV（表头与导入兼容，可再次导入）。"""
    import csv, io as _io, datetime
    devices = _load_dev()
    buf = _io.StringIO()
    w = csv.writer(buf)
    w.writerow(["名称", "型号", "设备编号", "校准日期", "校准有效期止", "测试类型", "出厂编号", "设备厂家", "备注"])
    for d in devices:
        w.writerow([d.get("name", ""), d.get("model", ""), d.get("mgmt_no", ""),
                    d.get("cal_date", ""), d.get("cal_end", ""), d.get("test_type", ""),
                    d.get("factory_no", ""), d.get("maker", ""), d.get("remark", "")])
    data = ("﻿" + buf.getvalue()).encode("utf-8")   # BOM 便于 Excel 正确识别 UTF-8
    fn = "设备库备份_%s.csv" % datetime.date.today().strftime("%Y%m%d")
    return send_file(_io.BytesIO(data), as_attachment=True, download_name=fn, mimetype="text/csv")

@app.route("/api/devices/upsert", methods=["POST"])
def api_devices_upsert():
    """新增或修改一台设备。body: {old_mgmt_no?, name, model, mgmt_no, cal_date, factory_no, maker, test_type, remark}
    有效期止由 校准日期 自动算。管理编号唯一。"""
    d = request.get_json(force=True) or {}
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "设备名称不能为空"}), 200
    cal_date = _dot_date(d.get("cal_date") or "")
    rec = {
        "name": name,
        "model": _dev_clean_model(d.get("model") or ""),
        "mgmt_no": (d.get("mgmt_no") or "").strip(),
        "cal_date": cal_date,
        "cal_end": _dev_cal_end(cal_date),
        "factory_no": (d.get("factory_no") or "").strip(),
        "maker": (d.get("maker") or "").strip(),
        "test_type": (d.get("test_type") or "").replace("\n", " ").strip(),
        "remark": (d.get("remark") or "").strip(),
    }
    devices = _load_dev()
    old_key = (d.get("old_mgmt_no") or "").strip()
    new_key = _dev_key(rec)
    created = True
    # 编辑：先按 old_mgmt_no 找，找不到再按新键找
    hit = -1
    if old_key:
        for i, x in enumerate(devices):
            if (x.get("mgmt_no") or "").strip() == old_key:
                hit = i
                break
    if hit < 0:
        for i, x in enumerate(devices):
            if _dev_key(x) == new_key:
                hit = i
                break
    # 若管理编号改成了别的已存在设备的编号 -> 拒绝，避免撞键
    if rec["mgmt_no"]:
        for i, x in enumerate(devices):
            if i != hit and (x.get("mgmt_no") or "").strip() == rec["mgmt_no"]:
                return jsonify({"ok": False, "error": "管理编号「%s」已存在，不能重复。" % rec["mgmt_no"]}), 200
    if hit >= 0:
        devices[hit] = rec
        created = False
    else:
        devices.append(rec)
    devices.sort(key=lambda d: (d.get("name", ""), d.get("mgmt_no", "")))
    _save_dev(devices)
    return jsonify({"ok": True, "created": created, "count": len(devices)})

@app.route("/api/devices/delete", methods=["POST"])
def api_devices_delete():
    """删除一台设备。body: {mgmt_no} 或 {name, model, factory_no}。"""
    d = request.get_json(force=True) or {}
    mgmt_no = (d.get("mgmt_no") or "").strip()
    target = _dev_key(d)
    devices = _load_dev()
    kept = []
    removed = 0
    for x in devices:
        same = (mgmt_no and (x.get("mgmt_no") or "").strip() == mgmt_no) or \
               (not mgmt_no and _dev_key(x) == target)
        if same and removed == 0:
            removed += 1
            continue
        kept.append(x)
    _save_dev(kept)
    return jsonify({"ok": True, "removed": removed, "count": len(kept)})

@app.route("/api/next_report_no")
def api_next_report_no():
    """今天的下一个报告编号：YJ/SYBG-<今天yyyymmdd><3位序号>。
    序号 = 生成记录里今天同前缀编号的最大值 +1；今天还没生成过则 001。"""
    import datetime, re
    today = datetime.date.today().strftime("%Y%m%d")
    prefix = "YJ/SYBG-" + today
    pat = re.compile(r"^YJ/SYBG-" + today + r"(\d{3})$")
    maxseq = 0
    for rec in _load_genlog():
        m = pat.match((rec or {}).get("report_no", "") or "")
        if m:
            maxseq = max(maxseq, int(m.group(1)))
    return jsonify({"report_no": "%s%03d" % (prefix, maxseq + 1)})

@app.route("/api/import_form", methods=["POST"])
def api_import_form():
    """上传「试验申请单」PDF，解析首页信息，返回可回填的字段。"""
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "未收到文件"}), 400
    try:
        import form_parser
    except Exception as ex:
        return jsonify({"ok": False, "error": "解析组件缺失：%s" % ex}), 200
    try:
        fields = form_parser.parse_pdf(f.read())
    except Exception as ex:
        return jsonify({"ok": False,
            "error": "无法解析这个 PDF：%s。可能是扫描图片（无文字层）——那样无法自动识别。" % ex}), 200
    if not fields:
        return jsonify({"ok": False, "error": "没从 PDF 里认出任何字段。请确认是「试验申请单」原件（带文字，不是扫描图）。"}), 200
    return jsonify({"ok": True, "fields": fields})

@app.route("/api/upload", methods=["POST"])
def api_upload():
    name = request.form.get("project", "")
    if not name:
        return jsonify({"ok": False, "error": "缺少项目名"}), 400
    saved = []
    skipped = []
    imgdir = os.path.join(pdir(name), "images")
    OK_EXT = (".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp", ".tif", ".tiff")
    for f in request.files.getlist("files"):
        ext = os.path.splitext(f.filename)[1].lower() or ".jpg"
        if ext not in OK_EXT:
            skipped.append({"orig": f.filename, "reason": "不支持的格式 %s" % ext})
            continue
        fn = uuid.uuid4().hex + ext
        try:
            f.save(os.path.join(imgdir, fn))
            saved.append({"file": fn, "orig": f.filename})
        except Exception as ex:
            skipped.append({"orig": f.filename, "reason": "保存失败：%s" % ex})
    return jsonify({"ok": True, "files": saved, "skipped": skipped})

@app.route("/api/image/<name>/<fn>")
def api_image(name, fn):
    imgdir = os.path.join(pdir(name), "images")
    if not os.path.exists(os.path.join(imgdir, fn)):
        abort(404)
    return send_from_directory(imgdir, fn)

@app.route("/api/generate", methods=["POST"])
def api_generate():
    try:
        data = request.get_json(force=True)
        name = data.get("name") or "报告"
        imgdir = os.path.join(pdir(name), "images")
        # 解析每张图 file -> 绝对路径
        proj = {"info": data.get("info", {}), "tests": []}
        for t in data.get("tests", []):
            tt = dict(t)
            groups = []
            for g in t.get("image_groups", []):
                imgs = []
                for im in g.get("images", []):
                    fp = os.path.join(imgdir, im.get("file", ""))
                    if os.path.exists(fp):
                        imgs.append({"path": fp, "caption": im.get("caption", "")})
                groups.append({"title": g.get("title", ""), "images": imgs})
            tt["image_groups"] = groups
            # 试验条件配图
            cimgs = []
            for im in t.get("condition_images", []):
                fp = os.path.join(imgdir, im.get("file", ""))
                if os.path.exists(fp):
                    cimgs.append({"path": fp, "caption": im.get("caption", "")})
            tt["condition_images"] = cimgs
            proj["tests"].append(tt)
        out = os.path.join(OUT_DIR, safe_name(name) + ".docx")
        try:
            E.generate(proj, out)
        except PermissionError:
            return jsonify({"ok": False,
                "error": "无法写入报告文件，可能它正在 WPS/Word 中打开。请先关闭已打开的「%s.docx」再重试。" % safe_name(name)}), 200
        # 记录本次生成的报告编号，供报告编号自动递增
        _append_genlog((data.get("info", {}) or {}).get("report_no", ""))
        return jsonify({"ok": True, "file": os.path.basename(out), "size": os.path.getsize(out)})
    except Exception as ex:
        import traceback
        tb = traceback.format_exc()
        try:
            with open(os.path.join(BASE, "构建", "错误日志.txt"), "w", encoding="utf-8") as f:
                f.write(tb)
        except Exception:
            pass
        return jsonify({"ok": False, "error": str(ex)}), 200

@app.route("/api/download")
def api_download():
    fn = request.args.get("file", "")
    fp = os.path.join(OUT_DIR, os.path.basename(fn))
    if not os.path.exists(fp):
        abort(404)
    return send_file(fp, as_attachment=True, download_name=fn)

@app.route("/api/open", methods=["POST"])
def api_open():
    """用默认程序(WPS/Word)打开生成的报告。仅当请求来自服务器本机时有效；
    远程浏览器访问时无法打开对方电脑上的 WPS，前端应改为下载文件。"""
    data = request.get_json(force=True)
    fn = os.path.basename(data.get("file", ""))
    fp = os.path.join(OUT_DIR, fn)
    if not os.path.exists(fp):
        return jsonify({"ok": False, "error": "文件不存在"}), 404
    if not _is_local_request():
        # 服务器无法在你的电脑上启动 WPS，交给前端下载后由本机 WPS 打开
        return jsonify({"ok": False, "remote": True,
            "error": "服务器无法打开你电脑上的 WPS，请改用“下载”后用本机 WPS 打开。"}), 200
    try:
        _native_open(fp)
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 200

@app.route("/api/open_folder", methods=["POST"])
def api_open_folder():
    """打开服务器本机输出文件夹并选中该文件（跨平台）。远程访问无效。"""
    data = request.get_json(force=True)
    fn = os.path.basename(data.get("file", ""))
    fp = os.path.join(OUT_DIR, fn)
    if not _is_local_request():
        return jsonify({"ok": False, "remote": True,
            "error": "服务器无法打开你电脑上的文件夹，请改用“下载”。"}), 200
    try:
        if os.path.exists(fp) and sys.platform.startswith("win"):
            os.system('explorer /select,"%s"' % fp)
        else:
            # Linux/macOS 或文件不存在时，打开输出目录
            _native_open(OUT_DIR)
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 200

PORT = 8731

def _open_browser():
    webbrowser.open("http://127.0.0.1:%d/" % PORT)

if __name__ == "__main__":
    threading.Timer(1.2, _open_browser).start()
    print("=" * 50)
    print("  MINEYE报告编辑助手已启动")
    print("  浏览器地址: http://127.0.0.1:%d/" % PORT)
    print("  ★ 使用期间请勿关闭本窗口 ★")
    print("=" * 50)
    app.run(host="127.0.0.1", port=PORT, debug=False, threaded=True)
