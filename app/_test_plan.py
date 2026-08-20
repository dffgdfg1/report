# -*- coding: utf-8 -*-
import importlib, plan_engine
importlib.reload(plan_engine)

IMG = r"C:\Users\admin\Desktop\报告生成器\模板库\std_images\65818967ab834f8186f21c261962e72c.png"

long_cond = ("工作模式：1.1；\n试件数量：1；\n试样跌落次数：2；\n跌落高度：1m；\n"
             "接受坠物的表面：混凝土地面；\n跌落方向：每个试样X，Y，Z方向，每次跌落沿一个"
             "方向（即用壳体的一侧照准地面)每次坠地后都进行外观检查。")
proj = {
    "info": {"sample_name": "DMS", "applicant": "潘英明"},
    "tests": [
        {"title": "低温工作", "standard": "Q/SQR T8-20-2019 7.1.2",
         "condition": "工作模式：3.2；\n试验时间：96h；\n试验温度：TLO=-40℃。",
         "requirement": "功能状态：A", "sample_no": "1#", "samples": [{"no": "1#"}],
         "overall_result": "合格",
         "condition_images": [{"path": IMG, "caption": ""}, {"path": IMG, "caption": ""}]},
        {"title": "跌落试验", "standard": "Q/SQR T8-20-2019 6.3",
         "condition": long_cond, "requirement": "外壳无开裂", "sample_no": "1#",
         "samples": [{"no": "1#"}], "overall_result": "合格",
         "condition_images": [{"path": IMG, "caption": ""}]},
    ],
}
out = r"C:\Users\admin\Desktop\报告生成器\输出\_测试3_试验计划.xlsx"
p = plan_engine.generate_plan(proj, out)

import openpyxl
from openpyxl.utils.units import EMU_to_pixels
wb = openpyxl.load_workbook(p)
ws = wb["试验计划"]
print("=== 计划表 图片行高与图片位置 ===")
for r in range(6, ws.max_row+1):
    h = ws.row_dimensions[r].height
    d = ws.cell(r,4)
    if d.value:
        print(f"行{r}: D文字={repr(d.value)[:30]} 行高={h} vAlign={d.alignment.vertical}")
print("\n=== 图片 (row 0-based, y偏移px, 尺寸px) ===")
for img in ws._images:
    a = img.anchor
    yoff = EMU_to_pixels(a._from.rowOff)
    w = EMU_to_pixels(a.ext.cx); h = EMU_to_pixels(a.ext.cy)
    print(f"  行{a._from.row+1}(1-based) col={a._from.col} y偏移={yoff}px 尺寸={w}x{h}px 底部={yoff+h}px")
