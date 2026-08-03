# -*- coding: utf-8 -*-
"""生成「标准条件库」Excel 录入模板 -> 模板库/标准库模板.xlsx
一行 = 一个(测试项目 × 车厂)的条件。填完用导入脚本转成 标准库.json。"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(BASE, "模板库", "标准库模板.xlsx")

# (列名, 列宽, 表头批注)
# 注：环境条件不进模板，实验室默认 18℃-28℃、25%RH-75%RH
COLS = [
    ("测试项目", 16, "必填。同名归到同一测试项目下，如：振动测试"),
    ("车厂", 12, "必填。如：比亚迪/蔚来；不分车厂就填：通用"),
    ("标准号/条款号", 26, "选填。如 ISO 16750-3:2012 4.1.2.4"),
    ("试验条件", 46, "选填。核心条件文字，换行按 Alt+Enter"),
    ("试验要求", 30, "选填。如 功能等级A。"),
    ("备注", 16, "选填。仅自己看，不进报告"),
]

# 示例行（灰字提示，可整行删掉）
SAMPLES = [
    ["振动测试", "比亚迪", "BYD-Q-SC 001 5.2", "扫频 10-500Hz，3轴各2h", "功能等级A。", "示例，可删"],
    ["振动测试", "蔚来", "NIO-TS-018 6.1", "随机振动 PSD谱见附表，3轴各8h", "功能等级A。", "示例，可删"],
    ["高温工作", "通用", "ISO 16750-4 5.1.2", "高温工作16h，通电监测", "功能等级A。", "示例，可删"],
]
HEAD_FILL = PatternFill("solid", fgColor="2F5496")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
HINT_FONT = Font(color="808080", italic=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "标准条件"
    # 表头
    for c, (name, width, note) in enumerate(COLS, 1):
        cell = ws.cell(1, c, name)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = width
        cell.comment = None
    ws.row_dimensions[1].height = 26
    # 表头批注（悬停显示填写说明）
    from openpyxl.comments import Comment
    for c, (name, width, note) in enumerate(COLS, 1):
        ws.cell(1, c).comment = Comment(note, "模板")
    # 示例行
    r = 2
    for row in SAMPLES:
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.font = HINT_FONT
            cell.alignment = WRAP
            cell.border = BORDER
        r += 1
    # 预留空白行，方便直接填
    for _ in range(200):
        for c in range(1, len(COLS) + 1):
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).alignment = WRAP
        r += 1
    ws.freeze_panes = "A2"  # 冻结表头
    # 说明页
    doc = wb.create_sheet("填写说明")
    tips = [
        "【怎么填】",
        "1. 一行 = 一个测试项目在某个车厂下的条件。",
        "2. 「测试项目」+「车厂」两列是关键：同一个测试项目名，写多行、车厂不同，",
        "   就会归到同一测试项目下的不同车厂，互不干扰。",
        "3. 不区分车厂的通用条件，「车厂」列填「通用」。",
        "4. 「试验条件」里要换行：按 Alt+Enter。",
        "5. 前 3 行是示例（灰字），填之前整行删掉即可。",
        "",
        "【填完之后】",
        "把本文件填好保存，回到软件里点『导入标准库』（或让维护者跑导入脚本），",
        "程序会自动读入并生成 标准库.json。重复导入同一 测试项目+车厂 会覆盖旧值。",
    ]
    for i, line in enumerate(tips, 1):
        doc.cell(i, 1, line)
    doc.column_dimensions["A"].width = 70
    wb.save(OUT)
    print("已生成模板:", OUT)


if __name__ == "__main__":
    build()
