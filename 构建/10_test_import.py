# -*- coding: utf-8 -*-
"""端到端自测：用模板自带示例行造一个已填表，调后端解析，验证 标准库.json 结构。"""
import os, io, sys, json
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "app"))
from openpyxl import load_workbook, Workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 造一个"用户已填好"的表：复制模板表头 + 3 行真实数据（把示例备注去掉）
tpl = load_workbook(os.path.join(BASE, "模板库", "标准库模板.xlsx"))["标准条件"]
header = [c.value for c in tpl[1]]
wb = Workbook(); ws = wb.active; ws.title = "标准条件"
ws.append(header)
ws.append(["振动测试", "比亚迪", "BYD 5.2", "扫频10-500Hz 3轴各2h", "功能等级A。", ""])
ws.append(["振动测试", "蔚来",   "NIO 6.1", "随机振动 3轴各8h", "功能等级A。", ""])
ws.append(["高温工作", "通用",   "ISO 16750-4", "85℃工作16h", "功能等级A。", ""])
ws.append(["", "", "", "", "", ""])  # 空行，应被跳过
buf = io.BytesIO(); wb.save(buf); buf.seek(0)

# 模拟后端解析逻辑（与 app.py 一致）
import app as A
A.STD_FILE = os.path.join(BASE, "模板库", "_test_标准库.json")
if os.path.exists(A.STD_FILE): os.remove(A.STD_FILE)

app = A.app
with app.test_client() as c:
    r = c.post("/api/standards/import", data={"file": (buf, "t.xlsx")},
               content_type="multipart/form-data")
    print("导入返回:", r.get_json())
    r2 = c.get("/api/standards")
    std = r2.get_json()

# 校验
assert set(std.keys()) == {"振动测试", "高温工作"}, "测试项目分组错误"
assert set(std["振动测试"].keys()) == {"比亚迪", "蔚来"}, "车厂分组错误"
assert std["振动测试"]["比亚迪"]["condition"] == "扫频10-500Hz 3轴各2h"
assert std["高温工作"]["通用"]["standard"] == "ISO 16750-4"
print("结构校验通过:")
print(json.dumps(std, ensure_ascii=False, indent=2))
os.remove(A.STD_FILE)
print("\n全部通过 ✓")
